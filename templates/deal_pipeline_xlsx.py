#!/usr/bin/env python3
"""
Deal Pipeline / Target Tracker — Excel template generator

The "deal list" deliverable: a multi-tab acquisition pipeline workbook of the kind every PE
deal team and searcher maintains. Tabs: Dashboard, Active Targets, Passed, Contacts.
Columns follow standard M&A pipeline convention (stage, closing probability, fit score, owner,
next step) adapted to the search-fund buy box.

Used by the /excel and /analyze-industry commands.

Usage:
    # Generate a ready-to-fill template (headers, dropdowns, example row):
    python3 templates/deal_pipeline_xlsx.py --output path/to/deal-pipeline.xlsx

    # Populate from a JSON data file:
    python3 templates/deal_pipeline_xlsx.py --data targets.json --output path/to/deal-pipeline.xlsx

JSON shape (all keys optional):
    {
      "vertical": "Home Health", "geography": "Rhode Island",
      "active":   [{"company": "...", "segment": "...", "location": "...",
                    "revenue": 4.2, "ebitda": 0.8, "recurring": 70, "concentration": 12,
                    "owner_transition": "Yes", "fit_score": 31,
                    "source": "BizBuySell", "stage": "Contacted", "probability": 20,
                    "lead": "DD", "last_contact": "2026-05-20", "next_step": "Send NDA",
                    "notes": "..."}],
      "passed":   [{"company": "...", "segment": "...", "reason": "...", "date": "..."}],
      "contacts": [{"name": "...", "role": "Broker", "company": "...", "type": "Broker",
                    "email": "...", "phone": "...", "last_contact": "...", "notes": "..."}]
    }
"""
import argparse
import json
import os

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

# ---- palette (matches the deck template's dark-blue / accent scheme) -------------
DARK_BLUE = "1B2A4A"
ACCENT_BLUE = "2E75B6"
LIGHT_GRAY = "F2F2F2"
ROW_ALT = "EBF5FB"
GREEN = "E8F8F5"
RED = "FDEDEC"
AMBER = "FEF9E7"
WHITE = "FFFFFF"

STAGES = ["Sourced", "Contacted", "NDA", "IOI", "LOI", "Diligence", "Closed", "Passed"]
YESNO = ["Yes", "No", "Partial", "Unknown"]

HEADER_FONT = Font(bold=True, color=WHITE, name="Calibri", size=11)
HEADER_FILL = PatternFill("solid", fgColor=DARK_BLUE)
TITLE_FONT = Font(bold=True, color=DARK_BLUE, name="Calibri", size=16)
SUB_FONT = Font(italic=True, color="7F8C8D", name="Calibri", size=10)
THIN = Side(style="thin", color="D5DBDB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# ---- column specs: (header, key, width, number_format) ---------------------------
ACTIVE_COLS = [
    ("#", None, 4, None),
    ("Company", "company", 28, None),
    ("Segment", "segment", 18, None),
    ("Location", "location", 18, None),
    ("Revenue ($M)", "revenue", 12, "#,##0.0"),
    ("EBITDA ($M)", "ebitda", 12, "#,##0.0"),
    ("Margin %", None, 10, "0%"),          # formula: EBITDA/Revenue
    ("Recurring %", "recurring", 11, "0\\%"),
    ("Top Cust %", "concentration", 10, "0\\%"),
    ("Owner Transition", "owner_transition", 16, None),
    ("Fit (/40)", "fit_score", 9, "0"),
    ("Source / Intermediary", "source", 22, None),
    ("Stage", "stage", 13, None),
    ("Prob %", "probability", 8, "0\\%"),
    ("Lead", "lead", 8, None),
    ("Last Contact", "last_contact", 13, None),
    ("Next Step", "next_step", 26, None),
    ("Notes", "notes", 34, None),
]


def _title(ws, title, subtitle):
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws["A2"] = subtitle
    ws["A2"].font = SUB_FONT


def _header_row(ws, headers, row):
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def build_active(ws, rows):
    _title(ws, "Active Targets", "Live acquisition pipeline — one row per target")
    hrow = 4
    headers = [c[0] for c in ACTIVE_COLS]
    _header_row(ws, headers, hrow)

    n = max(len(rows), 12)  # leave blank rows to fill in
    for i in range(n):
        r = hrow + 1 + i
        data = rows[i] if i < len(rows) else {}
        for c, (header, key, _w, fmt) in enumerate(ACTIVE_COLS, start=1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center",
                                       horizontal="left" if c in (2, 3, 12, 17, 18) else "center",
                                       wrap_text=c in (17, 18))
            if header == "#":
                cell.value = i + 1
            elif header == "Margin %":
                cell.value = f"=IF(E{r}=0,\"\",F{r}/E{r})"
            elif key and key in data:
                cell.value = data[key]
            if fmt:
                cell.number_format = fmt
            if (i % 2) == 1:
                cell.fill = PatternFill("solid", fgColor=ROW_ALT)

    last = hrow + n
    ws.freeze_panes = f"B{hrow + 1}"
    ws.auto_filter.ref = f"A{hrow}:{get_column_letter(len(headers))}{last}"
    for c, (_h, _k, w, _f) in enumerate(ACTIVE_COLS, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w

    # Dropdowns: Stage (col M=13) and Owner Transition (col J=10)
    stage_dv = DataValidation(type="list", formula1='"%s"' % ",".join(STAGES), allow_blank=True)
    yesno_dv = DataValidation(type="list", formula1='"%s"' % ",".join(YESNO), allow_blank=True)
    ws.add_data_validation(stage_dv)
    ws.add_data_validation(yesno_dv)
    stage_dv.add(f"M{hrow + 1}:M{last}")
    yesno_dv.add(f"J{hrow + 1}:J{last}")

    # Conditional formatting on Stage cells
    rng = f"M{hrow + 1}:M{last}"
    for stage, color in [("LOI", GREEN), ("Diligence", GREEN), ("Closed", GREEN),
                         ("Passed", RED), ("IOI", AMBER), ("NDA", AMBER)]:
        ws.conditional_formatting.add(
            rng, CellIsRule(operator="equal", formula=[f'"{stage}"'],
                            fill=PatternFill("solid", fgColor=color)))
    return last


def build_passed(ws, rows):
    _title(ws, "Passed / Abandoned", "Targets reviewed and declined — keep for institutional memory")
    headers = ["Company", "Segment", "Reason Passed", "Date"]
    widths = [28, 18, 44, 14]
    hrow = 4
    _header_row(ws, headers, hrow)
    for i in range(max(len(rows), 6)):
        r = hrow + 1 + i
        d = rows[i] if i < len(rows) else {}
        for c, (h, key) in enumerate(zip(headers, ["company", "segment", "reason", "date"]), start=1):
            cell = ws.cell(row=r, column=c, value=d.get(key))
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=(c == 3))
            if (i % 2) == 1:
                cell.fill = PatternFill("solid", fgColor=ROW_ALT)
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = f"A{hrow + 1}"


def build_contacts(ws, rows):
    _title(ws, "Contacts", "Brokers, advisors, and owners across the pipeline")
    headers = ["Name", "Role", "Company", "Type", "Email", "Phone", "Last Contact", "Notes"]
    keys = ["name", "role", "company", "type", "email", "phone", "last_contact", "notes"]
    widths = [22, 18, 24, 14, 26, 16, 13, 30]
    hrow = 4
    _header_row(ws, headers, hrow)
    typ_dv = DataValidation(type="list",
                            formula1='"Broker,M&A Advisor,Owner,Lender,Advisor,Other"', allow_blank=True)
    ws.add_data_validation(typ_dv)
    for i in range(max(len(rows), 8)):
        r = hrow + 1 + i
        d = rows[i] if i < len(rows) else {}
        for c, key in enumerate(keys, start=1):
            cell = ws.cell(row=r, column=c, value=d.get(key))
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=(c == 8))
            if (i % 2) == 1:
                cell.fill = PatternFill("solid", fgColor=ROW_ALT)
    typ_dv.add(f"D{hrow + 1}:D{hrow + max(len(rows), 8)}")
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = f"A{hrow + 1}"


def build_dashboard(ws, vertical, geography, active_last_row):
    _title(ws, f"Deal Pipeline — {vertical or '<vertical>'}",
           f"{geography or '<geography>'} · generated by analyze")
    ws["A4"] = "Pipeline by Stage"
    ws["A4"].font = Font(bold=True, color=DARK_BLUE, size=12)
    a = "'Active Targets'"
    rng = f"{a}!M5:M{active_last_row}"
    r = 5
    _header_row(ws, ["Stage", "Count"], r)
    for s in STAGES:
        r += 1
        ws.cell(row=r, column=1, value=s).border = BORDER
        cnt = ws.cell(row=r, column=2, value=f'=COUNTIF({rng},"{s}")')
        cnt.border = BORDER
        cnt.alignment = Alignment(horizontal="center")
    ws.cell(row=r + 1, column=1, value="Total active").font = Font(bold=True)
    ws.cell(row=r + 1, column=2,
            value=f'=COUNTA({a}!B5:B{active_last_row})').font = Font(bold=True)

    # Weighted pipeline value (Σ EBITDA × probability), informational
    ws["D4"] = "Snapshot"
    ws["D4"].font = Font(bold=True, color=DARK_BLUE, size=12)
    ws["D5"] = "Avg fit score (/40)"
    ws["E5"] = f'=IFERROR(AVERAGE({a}!K5:K{active_last_row}),0)'
    ws["D6"] = "Total EBITDA in pipeline ($M)"
    ws["E6"] = f'=SUM({a}!F5:F{active_last_row})'
    ws["D7"] = "Prob-weighted EBITDA ($M)"
    ws["E7"] = f'=SUMPRODUCT({a}!F5:F{active_last_row},{a}!N5:N{active_last_row})'
    for cell in ("E5", "E6", "E7"):
        ws[cell].number_format = "#,##0.0"
    for c, w in [("A", 18), ("B", 10), ("C", 3), ("D", 30), ("E", 12)]:
        ws.column_dimensions[c].width = w


def main():
    ap = argparse.ArgumentParser(description="Generate a deal-pipeline / target-tracker workbook.")
    ap.add_argument("--output", required=True, help="Output .xlsx path.")
    ap.add_argument("--data", help="Optional JSON data file (see module docstring).")
    args = ap.parse_args()

    d = {}
    if args.data:
        with open(args.data) as f:
            d = json.load(f)

    wb = Workbook()
    dash = wb.active
    dash.title = "Dashboard"
    active = wb.create_sheet("Active Targets")
    passed = wb.create_sheet("Passed")
    contacts = wb.create_sheet("Contacts")

    last = build_active(active, d.get("active", []))
    build_passed(passed, d.get("passed", []))
    build_contacts(contacts, d.get("contacts", []))
    build_dashboard(dash, d.get("vertical"), d.get("geography"), last)

    os.makedirs(os.path.dirname(os.path.abspath(args.output)), exist_ok=True)
    wb.save(args.output)
    print(f"Saved: {args.output}  (tabs: Dashboard, Active Targets, Passed, Contacts)")


if __name__ == "__main__":
    main()
